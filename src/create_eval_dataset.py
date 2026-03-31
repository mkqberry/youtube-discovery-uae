"""
Evaluation Dataset Creator
--------------------------
Reads a JSONL evaluation file and creates:
1. auto_eval_youtube/ folder with copied audio files (as index structure)
2. metadata.xlsx with columns: Audio_Filename, LLM_Generated_Transcript, Golden_Transcript, Remarks

Usage:
    python create_eval_dataset.py \
        --jsonl /path/to/parsed_evaluations_merged_file.jsonl \
        --audio_root /data/asr/youtube \
        --output_dir ./auto_eval_youtube
"""

import json
import os
import shutil
import random
import argparse
from collections import defaultdict
from pathlib import Path

import pandas as pd


def load_jsonl(path):
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


def sample_balanced(records, n_per_class=100, seed=42):
    """
    Sample n_per_class from overall_pass=True and n_per_class from overall_pass=False,
    trying to include at least one sample from each video_id.
    """
    random.seed(seed)

    # Skip records without evaluation key (parse_error records)
    valid = [r for r in records if "evaluation" in r]
    print(f"Valid records: {len(valid)} / {len(records)} (skipped {len(records)-len(valid)} parse errors)")

    # Separate by class
    true_records = [r for r in valid if r["evaluation"].get("overall_pass") is True]
    false_records = [r for r in valid if r["evaluation"].get("overall_pass") is False]

    def sample_with_video_coverage(pool, n):
        # Group by video_id (first part of index before '/')
        by_video = defaultdict(list)
        for r in pool:
            vid = r["index"].split("/")[0]
            by_video[vid].append(r)

        selected = []
        used_indices = set()

        # First: pick one from each video
        for vid, recs in by_video.items():
            chosen = random.choice(recs)
            selected.append(chosen)
            used_indices.add(chosen["index"])

        # If we already have more than n, randomly trim
        if len(selected) >= n:
            selected = random.sample(selected, n)
            return selected

        # Fill remaining from the full pool
        remaining_pool = [r for r in pool if r["index"] not in used_indices]
        random.shuffle(remaining_pool)
        needed = n - len(selected)
        selected.extend(remaining_pool[:needed])

        return selected

    true_sample = sample_with_video_coverage(true_records, n_per_class)
    false_sample = sample_with_video_coverage(false_records, n_per_class)

    print(f"✅ Sampled {len(true_sample)} overall_pass=True records")
    print(f"❌ Sampled {len(false_sample)} overall_pass=False records")

    return true_sample + false_sample


def copy_audio_files(sampled, audio_root, output_dir):
    """
    Copy audio files maintaining index structure into output_dir/audio/
    Returns list of relative audio paths (or None if file not found).
    """
    audio_out = Path(output_dir) / "audio"
    audio_out.mkdir(parents=True, exist_ok=True)

    audio_filenames = []

    for record in sampled:
        index = record["index"]
        # Audio files are stored under wavs/ subfolder
        # index: video_id/segment_folder/segment_00000.wav
        # actual path: audio_root/video_id/segment_folder/wavs/segment_00000.wav
        parts = Path(index).parts  # (video_id, segment_folder, segment_00000.wav)
        if len(parts) == 3:
            src_path = Path(audio_root) / parts[0] / parts[1] / "wavs" / parts[2]
        else:
            src_path = Path(audio_root) / index
        rel_path = Path(index)

        # Destination mirrors the index structure
        dst_path = audio_out / rel_path
        dst_path.parent.mkdir(parents=True, exist_ok=True)

        if src_path.exists():
            shutil.copy2(src_path, dst_path)
            audio_filenames.append(str(rel_path))
        else:
            print(f"⚠️  Audio not found: {src_path}")
            audio_filenames.append(str(rel_path) + "  [FILE NOT FOUND]")

    return audio_filenames


def create_metadata_xlsx(sampled, audio_filenames, output_dir):
    rows = []
    for record, audio_fn in zip(sampled, audio_filenames):
        eval_data = record.get("evaluation", {})
        overall_pass = eval_data.get("overall_pass", None)
        rows.append({
            "Audio_Filename": audio_fn,
            "Overall_Pass": overall_pass,
            "LLM_Generated_Transcript": record.get("transcript", ""),
            "Golden_Transcript": "",
            "Remarks": "",
        })

    df = pd.DataFrame(rows, columns=["Audio_Filename", "Overall_Pass", "LLM_Generated_Transcript", "Golden_Transcript", "Remarks"])

    # Sort: True first, then False for readability
    df = df.sort_values("Overall_Pass", ascending=False).reset_index(drop=True)

    out_path = Path(output_dir) / "metadata.xlsx"

    # Color coding with openpyxl
    import openpyxl
    from openpyxl.styles import PatternFill

    df.to_excel(out_path, index=False)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    pass_col_idx = df.columns.get_loc("Overall_Pass") + 1  # openpyxl is 1-indexed

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[pass_col_idx - 1]
        if cell.value is True:
            cell.fill = green_fill
        elif cell.value is False:
            cell.fill = red_fill

    wb.save(out_path)
    print(f"📊 metadata.xlsx saved to: {out_path}")
    return out_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--jsonl", required=True, help="Path to merged JSONL evaluation file")
    parser.add_argument("--audio_root", required=True, help="Root directory of audio files (e.g. /data/asr/youtube)")
    parser.add_argument("--output_dir", default="./auto_eval_youtube", help="Output directory")
    parser.add_argument("--n_per_class", type=int, default=100, help="Number of samples per class (default: 100)")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    args = parser.parse_args()

    print(f"📂 Loading JSONL: {args.jsonl}")
    records = load_jsonl(args.jsonl)
    print(f"   Total records: {len(records)}")

    print("\n🎯 Sampling balanced dataset...")
    sampled = sample_balanced(records, n_per_class=args.n_per_class, seed=args.seed)
    print(f"   Total sampled: {len(sampled)}")

    # Shuffle combined
    random.seed(args.seed)
    random.shuffle(sampled)

    print(f"\n📁 Copying audio files to: {args.output_dir}/audio/")
    audio_filenames = copy_audio_files(sampled, args.audio_root, args.output_dir)

    print(f"\n📝 Creating metadata.xlsx...")
    create_metadata_xlsx(sampled, audio_filenames, args.output_dir)

    print("\n✅ Done!")
    print(f"   Output directory: {args.output_dir}/")
    print(f"   ├── audio/          (copied wav files mirroring index structure)")
    print(f"   └── metadata.xlsx   (Audio_Filename, LLM_Generated_Transcript, Golden_Transcript, Remarks)")


if __name__ == "__main__":
    main()
